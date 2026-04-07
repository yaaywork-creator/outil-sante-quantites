import io
import math
import random
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.express as px
import psycopg2
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="EDDAQAQ EXPERTISES",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

APP_PASSWORD = st.secrets.get("APP_PASSWORD", "EDDAQAQ2026")
DATABASE_URL = st.secrets.get("DATABASE_URL", "")

MONTHS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]

YEARS = ["Année 1", "Année 2", "Année 3", "Année 4", "Année 5"]

DEFAULT_CA = {
    "Année 1": 4_000_000.0,
    "Année 2": 5_000_000.0,
    "Année 3": 6_000_000.0,
    "Année 4": 7_200_000.0,
    "Année 5": 7_920_000.0,
}

DEFAULT_SECTIONS = {
    "Clinique": [
        ("Hospitalisation Normale", 1150),
        ("Hôpital de jour", 800),
        ("Soins intensifs", 2000),
        ("Réanimation", 3000),
        ("Lits d’urgences", 800),
        ("Salle de réveil", 600),
        ("Chambre individuelle", 1500),
        ("Consultation générale", 250),
        ("Consultation spécialisée", 400),
        ("Petite chirurgie", 900),
        ("Acte infirmier", 120),
        ("Pansement spécialisé", 140),
    ],
    "Laboratoire": [
        ("NFS", 80),
        ("CRP", 90),
        ("Glycémie", 40),
        ("HbA1c", 120),
        ("Bilan hépatique", 180),
        ("Bilan rénal", 150),
        ("Ionogramme", 130),
        ("TSH", 140),
        ("Ferritine", 160),
        ("Vitamine D", 220),
        ("ECBU", 70),
        ("Bilan lipidique", 150),
        ("Sérologie", 190),
    ],
    "Centre de radiologie": [
        ("RX PHOTO", 50),
        ("RX(face)", 200),
        ("RX face et profil", 300),
        ("Echo Abdo", 500),
        ("Echo parties molles", 600),
        ("Echo cervical", 600),
        ("Echo mammaire", 600),
        ("Mammo", 600),
        ("Mammo+Echo", 800),
        ("Scanner cérébral", 1500),
        ("Scanner abdominal", 1500),
        ("TDM FACE", 1500),
        ("TDM SINUS", 1500),
        ("TDM cervical/dorsal/lombaire", 1500),
        ("TDM bassin", 1500),
        ("TDM hanche/genou/épaule/coude", 1500),
        ("Scanner thoracique", 1500),
        ("TDM rocher", 2000),
        ("Scanner abdomino pelvien", 2000),
        ("Scanner thoraco abdomino pelvien", 3000),
        ("Angio scanner", 3000),
    ],
}

# =========================================================
# STYLE
# =========================================================
st.markdown("""
<style>
:root{
    --bg1:#040b16;
    --bg2:#081220;
    --bg3:#0b1830;
    --card:#0d2037;
    --txt:#f4f8fd;
}

html, body, [class*="css"] {
    font-family: "Segoe UI", Arial, sans-serif;
}

.stApp {
    background:
        radial-gradient(circle at top left, rgba(58,110,175,0.18), transparent 24%),
        radial-gradient(circle at top right, rgba(201,154,103,0.08), transparent 18%),
        linear-gradient(145deg, var(--bg1) 0%, var(--bg2) 40%, var(--bg3) 100%);
    color: white;
}

.block-container {
    padding-top: 1rem;
    padding-bottom: 2rem;
    max-width: 1500px;
}

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #08111f 0%, #0b1527 100%);
    border-right: 1px solid rgba(255,255,255,0.06);
}
section[data-testid="stSidebar"] * {
    color: #eef4fb !important;
}

.hero {
    background: linear-gradient(135deg, #08111f 0%, #12345c 60%, #1b4f86 100%);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 26px;
    padding: 24px 28px;
    box-shadow: 0 18px 42px rgba(0,0,0,0.34);
    margin-bottom: 18px;
}
.hero h1 {
    color: #ffffff;
    margin: 0 0 8px 0;
    font-size: 2.25rem;
    font-weight: 800;
}
.hero p {
    color: #e0ebf8;
    margin: 0;
    font-size: 1rem;
    line-height: 1.65;
}

.metric-box {
    background: linear-gradient(180deg, rgba(16,36,63,0.98), rgba(10,23,40,0.98));
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 20px;
    padding: 18px 20px;
    box-shadow: 0 12px 28px rgba(0,0,0,0.22);
}
.metric-label {
    color: #b7c8dd;
    font-size: 0.92rem;
}
.metric-value {
    color: #ffffff;
    font-size: 1.5rem;
    font-weight: 800;
    margin-top: 6px;
}

.card {
    background: linear-gradient(180deg, rgba(10,24,43,0.98), rgba(8,20,36,0.98));
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 22px;
    padding: 18px;
    box-shadow: 0 14px 32px rgba(0,0,0,0.22);
}

.section-title {
    color: #ffffff;
    font-weight: 800;
    font-size: 1.15rem;
    margin-bottom: 8px;
}

.small-text {
    color: #c7d5e7;
    font-size: 0.94rem;
    margin-bottom: 10px;
}

.badge {
    display: inline-block;
    padding: 6px 10px;
    border-radius: 999px;
    background: rgba(255,255,255,0.08);
    color: #e8f1fb;
    font-size: 0.85rem;
    margin-right: 6px;
}

div[data-testid="stDataFrame"] {
    background: rgba(255,255,255,0.98);
    border-radius: 16px;
    padding: 4px;
}
div[data-testid="stDataEditor"] {
    background: rgba(255,255,255,0.98);
    border-radius: 16px;
    padding: 4px;
}

.stTabs [data-baseweb="tab-list"] {
    gap: 10px;
}
.stTabs [data-baseweb="tab"] {
    background: rgba(255,255,255,0.08);
    border-radius: 14px;
    color: #dce8f7;
    padding: 10px 16px;
    border: 1px solid rgba(255,255,255,0.06);
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(180deg, #2c5d98 0%, #234b85 100%) !important;
    color: white !important;
    box-shadow: 0 8px 18px rgba(44,93,152,0.35);
}

.stButton > button,
.stDownloadButton > button {
    border-radius: 14px !important;
    border: 1px solid rgba(255,255,255,0.10) !important;
    font-weight: 700 !important;
    min-height: 46px !important;
    background: linear-gradient(180deg, #2c5d98 0%, #234b85 100%) !important;
    color: #ffffff !important;
    opacity: 1 !important;
}

.stButton > button:hover,
.stDownloadButton > button:hover {
    background: linear-gradient(180deg, #3773b9 0%, #295693 100%) !important;
    color: #ffffff !important;
}

.stButton > button:disabled,
.stDownloadButton > button:disabled {
    background: linear-gradient(180deg, #556479 0%, #4a566b 100%) !important;
    color: #ffffff !important;
    opacity: 1 !important;
}

.kpi-note {
    color: #d6e4f6;
    font-size: 0.95rem;
    margin-top: 8px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# DB
# =========================================================
def get_db_connection():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL manquant dans les secrets Streamlit.")
    return psycopg2.connect(DATABASE_URL)

def init_db():
    con = get_db_connection()
    cur = con.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS section_items (
        id SERIAL PRIMARY KEY,
        section_name TEXT NOT NULL,
        item_name TEXT NOT NULL,
        unit_price NUMERIC(18,2) NOT NULL,
        sort_order INTEGER NOT NULL DEFAULT 0
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS section_ca (
        id SERIAL PRIMARY KEY,
        section_name TEXT NOT NULL,
        year_name TEXT NOT NULL,
        ca_target NUMERIC(18,2) NOT NULL
    )
    """)

    cur.execute("SELECT COUNT(*) FROM section_items")
    count_items = cur.fetchone()[0]
    if count_items == 0:
        for section, items in DEFAULT_SECTIONS.items():
            for i, (name, price) in enumerate(items, start=1):
                cur.execute("""
                    INSERT INTO section_items (section_name, item_name, unit_price, sort_order)
                    VALUES (%s, %s, %s, %s)
                """, (section, name, float(price), i))

    cur.execute("SELECT COUNT(*) FROM section_ca")
    count_ca = cur.fetchone()[0]
    if count_ca == 0:
        for section in DEFAULT_SECTIONS.keys():
            for year_name, value in DEFAULT_CA.items():
                cur.execute("""
                    INSERT INTO section_ca (section_name, year_name, ca_target)
                    VALUES (%s, %s, %s)
                """, (section, year_name, float(value)))

    con.commit()
    cur.close()
    con.close()

def load_section_items(section_name: str) -> pd.DataFrame:
    con = get_db_connection()
    query = """
        SELECT item_name AS "Acte / Examen", unit_price AS "Prix Unitaire"
        FROM section_items
        WHERE section_name = %s
        ORDER BY sort_order, id
    """
    df = pd.read_sql(query, con, params=[section_name])
    con.close()

    if df.empty:
        df = pd.DataFrame(DEFAULT_SECTIONS[section_name], columns=["Acte / Examen", "Prix Unitaire"])

    df["Prix Unitaire"] = pd.to_numeric(df["Prix Unitaire"], errors="coerce").fillna(0.0)
    return df

def load_section_ca(section_name: str) -> dict:
    con = get_db_connection()
    cur = con.cursor()
    cur.execute("""
        SELECT year_name, ca_target
        FROM section_ca
        WHERE section_name = %s
    """, (section_name,))
    rows = cur.fetchall()
    cur.close()
    con.close()

    result = DEFAULT_CA.copy()
    for year_name, ca_target in rows:
        result[year_name] = float(ca_target)
    return result

def save_section_items(section_name: str, df: pd.DataFrame):
    con = get_db_connection()
    cur = con.cursor()
    cur.execute("DELETE FROM section_items WHERE section_name = %s", (section_name,))
    for idx, row in df.reset_index(drop=True).iterrows():
        item_name = str(row["Acte / Examen"]).strip()
        unit_price = float(row["Prix Unitaire"])
        if item_name:
            cur.execute("""
                INSERT INTO section_items (section_name, item_name, unit_price, sort_order)
                VALUES (%s, %s, %s, %s)
            """, (section_name, item_name, unit_price, idx + 1))
    con.commit()
    cur.close()
    con.close()

def save_section_ca(section_name: str, ca_dict: dict):
    con = get_db_connection()
    cur = con.cursor()
    cur.execute("DELETE FROM section_ca WHERE section_name = %s", (section_name,))
    for year_name, value in ca_dict.items():
        cur.execute("""
            INSERT INTO section_ca (section_name, year_name, ca_target)
            VALUES (%s, %s, %s)
        """, (section_name, year_name, float(value)))
    con.commit()
    cur.close()
    con.close()

# =========================================================
# AUTH
# =========================================================
def login_page():
    st.markdown("""
    <div class="hero">
        <h1>Plateforme EDDAQAQ EXPERTISES</h1>
        <p>Connexion sécurisée à l’outil santé avec sauvegarde durable, Excel premium et PDF structuré.</p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 1.1, 1])
    with c2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Connexion</div>', unsafe_allow_html=True)
        pwd = st.text_input("Mot de passe", type="password")
        if st.button("Se connecter", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect.")
        st.markdown('</div>', unsafe_allow_html=True)

# =========================================================
# HELPERS
# =========================================================
def clean_items_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Acte / Examen"] = df["Acte / Examen"].astype(str).str.strip()
    df["Prix Unitaire"] = pd.to_numeric(df["Prix Unitaire"], errors="coerce").fillna(0.0)
    df = df[df["Acte / Examen"] != ""].reset_index(drop=True)
    return df

def add_line_numbers(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy().reset_index(drop=True)
    out.insert(0, "N°", range(1, len(out) + 1))
    return out

def format_df_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if col not in ["Acte / Examen", "Mois", "Année", "N°"]:
            if pd.api.types.is_numeric_dtype(out[col]):
                out[col] = out[col].round(2)
    return out

def money(v):
    try:
        return f"{float(v):,.2f}".replace(",", " ")
    except Exception:
        return str(v)

def slugify_name(text: str) -> str:
    return (
        text.lower()
        .replace(" ", "_")
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
        .replace("î", "i")
        .replace("ï", "i")
        .replace("ç", "c")
        .replace("/", "_")
        .replace("'", "")
    )

def to_cents(x: float) -> int:
    return int(round(float(x) * 100))

def from_cents(x: int) -> float:
    return x / 100.0

def gcd_list(values):
    g = 0
    for v in values:
        g = math.gcd(g, int(v))
    return g

def monthly_seasonality(section_name: str):
    if section_name == "Clinique":
        arr = np.array([0.083, 0.078, 0.084, 0.085, 0.086, 0.084, 0.079, 0.077, 0.084, 0.086, 0.087, 0.087], dtype=float)
    elif section_name == "Laboratoire":
        arr = np.array([0.084, 0.079, 0.084, 0.083, 0.084, 0.083, 0.081, 0.079, 0.083, 0.086, 0.087, 0.087], dtype=float)
    else:
        arr = np.array([0.082, 0.078, 0.083, 0.084, 0.085, 0.084, 0.081, 0.079, 0.084, 0.086, 0.087, 0.087], dtype=float)
    return arr / arr.sum()

def get_item_weights(section_name: str, items_df: pd.DataFrame):
    weights = []
    names = items_df["Acte / Examen"].astype(str).tolist()

    if section_name == "Clinique":
        manual = {
            "Consultation générale": 1.8,
            "Consultation spécialisée": 1.4,
            "Acte infirmier": 1.8,
            "Pansement spécialisé": 1.5,
            "Hospitalisation Normale": 1.1,
            "Hôpital de jour": 1.2,
            "Lits d’urgences": 1.1,
            "Salle de réveil": 0.8,
            "Chambre individuelle": 0.7,
            "Petite chirurgie": 1.0,
            "Soins intensifs": 0.45,
            "Réanimation": 0.25,
        }
        for n in names:
            weights.append(manual.get(n, 1.0))
    elif section_name == "Laboratoire":
        manual = {
            "Glycémie": 1.8,
            "NFS": 1.8,
            "CRP": 1.4,
            "ECBU": 1.2,
            "HbA1c": 1.0,
            "Bilan rénal": 1.0,
            "Ionogramme": 0.9,
            "TSH": 0.8,
            "Ferritine": 0.75,
            "Vitamine D": 0.55,
            "Sérologie": 0.65,
            "Bilan hépatique": 0.8,
            "Bilan lipidique": 0.85,
        }
        for n in names:
            weights.append(manual.get(n, 1.0))
    else:
        manual = {
            "RX PHOTO": 1.7,
            "RX(face)": 1.3,
            "RX face et profil": 1.15,
            "Echo Abdo": 1.0,
            "Echo parties molles": 0.9,
            "Echo cervical": 0.85,
            "Echo mammaire": 0.85,
            "Mammo": 0.8,
            "Mammo+Echo": 0.7,
            "Scanner cérébral": 0.55,
            "Scanner abdominal": 0.55,
            "TDM FACE": 0.45,
            "TDM SINUS": 0.45,
            "TDM cervical/dorsal/lombaire": 0.40,
            "TDM bassin": 0.40,
            "TDM hanche/genou/épaule/coude": 0.38,
            "Scanner thoracique": 0.48,
            "TDM rocher": 0.25,
            "Scanner abdomino pelvien": 0.30,
            "Scanner thoraco abdomino pelvien": 0.18,
            "Angio scanner": 0.14,
        }
        for n in names:
            weights.append(manual.get(n, 1.0))

    return np.array(weights, dtype=float)

def find_exact_addition(diff_cents: int, prices_cents: list[int]):
    if diff_cents == 0:
        return [0] * len(prices_cents)

    prev = [-1] * (diff_cents + 1)
    used = [-1] * (diff_cents + 1)
    prev[0] = -2

    for amount in range(diff_cents + 1):
        if prev[amount] != -1:
            for idx, p in enumerate(prices_cents):
                nxt = amount + p
                if nxt <= diff_cents and prev[nxt] == -1:
                    prev[nxt] = amount
                    used[nxt] = idx

    if prev[diff_cents] == -1:
        return None

    out = [0] * len(prices_cents)
    cur = diff_cents
    while cur > 0:
        idx = used[cur]
        out[idx] += 1
        cur = prev[cur]
    return out

# =========================================================
# GENERATION
# =========================================================
def generate_realistic_year(
    section_name: str,
    items_df: pd.DataFrame,
    ca_target: float,
    seed: int = 42,
    allow_decimals: bool = False
):
    rng = np.random.default_rng(seed)
    df = items_df.copy().reset_index(drop=True)

    prices = df["Prix Unitaire"].astype(float).tolist()
    prices_cents = [to_cents(p) for p in prices]
    target_cents = to_cents(ca_target)

    if any(p <= 0 for p in prices_cents):
        raise ValueError("Tous les prix unitaires doivent être > 0.")

    weights = get_item_weights(section_name, df)

    price_factor = np.array([1 / max(p, 1) for p in prices], dtype=float)
    price_factor = price_factor / price_factor.mean()

    blended = weights * (price_factor ** 0.35)
    blended = blended / blended.sum()

    desired_ca = blended * ca_target

    qty = []
    for desired, price in zip(desired_ca, prices):
        q = int(math.floor(desired / price))
        qty.append(max(0, q))

    current_cents = sum(q * p for q, p in zip(qty, prices_cents))
    diff_cents = target_cents - current_cents

    if diff_cents < 0:
        raise ValueError("Erreur interne : la base dépasse le CA cible.")

    add = find_exact_addition(diff_cents, prices_cents)

    if add is None:
        solved = False
        order = list(np.argsort(prices_cents))
        for j in order:
            if qty[j] > 0:
                qty[j] -= 1
                new_diff = target_cents - sum(q * p for q, p in zip(qty, prices_cents))
                if new_diff >= 0:
                    add = find_exact_addition(new_diff, prices_cents)
                    if add is not None:
                        solved = True
                        break
                qty[j] += 1
        if not solved:
            if not allow_decimals:
                g = gcd_list(prices_cents)
                raise ValueError(
                    f"Impossible d’atteindre exactement le CA cible avec des quantités entières. "
                    f"Active l’option décimale. "
                    f"Le pas exact minimum entre combinaisons entières est de {g/100:.2f}."
                )

    if add is not None:
        qty = [q + a for q, a in zip(qty, add)]

    final_cents = sum(q * p for q, p in zip(qty, prices_cents))

    if final_cents != target_cents and not allow_decimals:
        raise ValueError(
            f"Le CA généré ({from_cents(final_cents):,.2f}) n’est pas exactement égal au CA cible ({ca_target:,.2f}). "
            f"Active l’option décimale si tu veux autoriser un léger ajustement."
        )

    rows = []

    for i, row in df.iterrows():
        item = row["Acte / Examen"]
        price = float(row["Prix Unitaire"])
        total_qty = float(qty[i])

        item_profile = monthly_seasonality(section_name) * rng.uniform(0.92, 1.08, size=12)
        item_profile = item_profile / item_profile.sum()

        if not allow_decimals:
            monthly_qty = rng.multinomial(int(round(total_qty)), item_profile).astype(float)
        else:
            monthly_qty = np.round(item_profile * total_qty, 2)
            gap = round(total_qty - float(monthly_qty.sum()), 2)
            if abs(gap) > 0:
                monthly_qty[np.argmax(item_profile)] += gap

        out = {
            "Acte / Examen": item,
            "Prix Unitaire": price
        }

        total_ca = 0.0
        total_qty_line = 0.0

        for m, q in zip(MONTHS, monthly_qty):
            q_value = round(float(q), 2) if allow_decimals else int(round(float(q)))
            ca = q_value * price
            out[f"Qté {m}"] = q_value
            out[f"CA {m}"] = float(ca)
            total_ca += ca
            total_qty_line += q_value

        out["Qté Totale"] = round(total_qty_line, 2) if allow_decimals else int(round(total_qty_line))
        out["CA Total"] = float(total_ca)
        rows.append(out)

    detail_df = pd.DataFrame(rows)

    total_generated = round(float(detail_df["CA Total"].sum()), 2)

    if allow_decimals and round(total_generated, 2) != round(ca_target, 2):
        diff = round(ca_target - total_generated, 2)
        if abs(diff) > 0:
            adjust_idx = int(np.argmax(blended / np.array(prices)))
            adjust_price = float(df.loc[adjust_idx, "Prix Unitaire"])
            adjustment_qty = round(diff / adjust_price, 2)

            detail_df.loc[adjust_idx, "Qté Décembre"] = round(float(detail_df.loc[adjust_idx, "Qté Décembre"]) + adjustment_qty, 2)
            detail_df.loc[adjust_idx, "CA Décembre"] = round(float(detail_df.loc[adjust_idx, "Qté Décembre"]) * adjust_price, 2)

            qty_cols = [f"Qté {m}" for m in MONTHS]
            ca_cols = [f"CA {m}" for m in MONTHS]

            detail_df.loc[adjust_idx, "Qté Totale"] = round(float(detail_df.loc[adjust_idx, qty_cols].sum()), 2)
            detail_df.loc[adjust_idx, "CA Total"] = round(float(detail_df.loc[adjust_idx, ca_cols].sum()), 2)

            total_generated = round(float(detail_df["CA Total"].sum()), 2)

    if round(total_generated, 2) != round(ca_target, 2):
        raise ValueError(
            f"Le CA généré ({total_generated:,.2f}) n'est pas égal au CA cible ({ca_target:,.2f})."
        )

    return detail_df, total_generated

def build_monthly_summary(detail_df):
    rows = []
    for month in MONTHS:
        qty_total = float(detail_df[f"Qté {month}"].sum())
        qty_total = int(round(qty_total)) if abs(qty_total - round(qty_total)) < 1e-9 else round(qty_total, 2)
        rows.append({
            "Mois": month,
            "Quantité Totale": qty_total,
            "CA Mensuel": float(detail_df[f"CA {month}"].sum())
        })
    return pd.DataFrame(rows)

# =========================================================
# EXCEL
# =========================================================
def excel_apply_base_style(ws, title, max_col):
    navy = PatternFill("solid", fgColor="081A32")
    blue = PatternFill("solid", fgColor="123A64")
    cream = PatternFill("solid", fgColor="F8F5EF")
    light_blue = PatternFill("solid", fgColor="EAF2FB")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    white_title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    white_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="000000")
    bold_font = Font(name="Calibri", size=10, bold=True, color="000000")

    thin = Side(style="thin", color="D9E1F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = navy
    cell.font = white_title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for c in ws[2]:
        c.fill = blue
        c.font = white_header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r in range(3, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = cream if r % 2 == 1 else light_blue
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    last_row_text_1 = str(ws.cell(ws.max_row, 1).value or "").lower()
    last_row_text_2 = str(ws.cell(ws.max_row, 2).value or "").lower()
    if "total" in last_row_text_1 or "total" in last_row_text_2:
        for c in range(1, max_col + 1):
            cell = ws.cell(ws.max_row, c)
            cell.fill = total_fill
            cell.font = bold_font
            cell.border = border

    ws.freeze_panes = "A3"

def excel_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 28)

def excel_format_numbers(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float, np.floating)):
                cell.number_format = '#,##0.00'

def export_excel_styled(section_name, ca_dict, all_results, all_monthly):
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "CA_Cible"
    ws0.append(["N°", "Année", "CA cible"])
    for i, y in enumerate(YEARS, start=1):
        ws0.append([i, y, float(ca_dict[y])])
    ws0.append(["", "TOTAL", float(sum(ca_dict.values()))])
    excel_apply_base_style(ws0, f"{section_name} - CA cible", 3)
    excel_format_numbers(ws0)
    excel_auto_width(ws0)

    ws_global = wb.create_sheet(title="Synthese_Globale")
    ws_global.append(["N°", "Année", "CA cible", "CA généré", "Écart"])
    for i, y in enumerate(YEARS, start=1):
        ca_gen = float(all_results[y]["CA Total"].sum())
        ws_global.append([i, y, float(ca_dict[y]), ca_gen, ca_gen - float(ca_dict[y])])
    ws_global.append(["", "TOTAL", float(sum(ca_dict.values())), float(sum(float(all_results[y]['CA Total'].sum()) for y in YEARS)), 0.0])
    excel_apply_base_style(ws_global, f"{section_name} - Synthèse globale", 5)
    excel_format_numbers(ws_global)
    excel_auto_width(ws_global)

    for y in YEARS:
        detail_df = add_line_numbers(all_results[y]).copy()
        monthly_df = add_line_numbers(all_monthly[y]).copy()

        ws1 = wb.create_sheet(title=y.replace(" ", "_")[:31])
        ws1.append(detail_df.columns.tolist())
        for row in detail_df.itertuples(index=False, name=None):
            ws1.append(list(row))
        ws1.append(
            ["", "TOTAL"] +
            [""] * (len(detail_df.columns) - 4) +
            [detail_df["Qté Totale"].sum(), float(detail_df["CA Total"].sum())]
        )
        excel_apply_base_style(ws1, f"{section_name} - {y} - Détail complet", len(detail_df.columns))
        excel_format_numbers(ws1)
        excel_auto_width(ws1)

        ws2 = wb.create_sheet(title=(y.replace(" ", "_") + "_Mensuel")[:31])
        ws2.append(monthly_df.columns.tolist())
        for row in monthly_df.itertuples(index=False, name=None):
            ws2.append(list(row))
        ws2.append(["", "TOTAL", monthly_df["Quantité Totale"].sum(), float(monthly_df["CA Mensuel"].sum())])
        excel_apply_base_style(ws2, f"{section_name} - {y} - Synthèse mensuelle", len(monthly_df.columns))
        excel_format_numbers(ws2)
        excel_auto_width(ws2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =========================================================
# PDF
# =========================================================
def dataframe_to_pdf_data(df):
    temp = df.copy()
    for col in temp.columns:
        if col not in ["N°", "Acte / Examen", "Mois", "Année"]:
            temp[col] = temp[col].apply(
                lambda x: f"{x:,.2f}".replace(",", " ") if isinstance(x, (int, float, np.floating)) else str(x)
            )
    return [temp.columns.tolist()] + temp.values.tolist()

def split_dataframe_for_pdf(df, max_rows=12):
    if df.empty:
        return [df]
    chunks = []
    for start in range(0, len(df), max_rows):
        chunks.append(df.iloc[start:start + max_rows].copy())
    return chunks

def build_pdf_table(df, header_color="#123A64", body_color="#F7F4EF", font_size=8.2, col_widths=None):
    data = dataframe_to_pdf_data(df)
    tbl = Table(data, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(body_color)),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9AA7B6")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tbl

def export_pdf(section_name, ca_dict, all_results, all_monthly):
    output = io.BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#081A32"),
        spaceAfter=8,
    )
    sub_style = ParagraphStyle(
        "sub_style",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor("#123A64"),
        spaceAfter=6,
        spaceBefore=6,
    )
    normal_style = ParagraphStyle(
        "normal_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#202B37")
    )
    small_style = ParagraphStyle(
        "small_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#475467")
    )

    elements = []
    elements.append(Paragraph("EDDAQAQ EXPERTISES", title_style))
    elements.append(Paragraph(f"Rapport structuré - {section_name}", sub_style))
    elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.25 * cm))

    ca_df = pd.DataFrame({
        "N°": range(1, len(YEARS) + 1),
        "Année": YEARS,
        "CA cible": [float(ca_dict[y]) for y in YEARS]
    })
    elements.append(Paragraph("1. CA cible par année", sub_style))
    elements.append(build_pdf_table(ca_df, header_color="#081A32", body_color="#F7F4EF", font_size=8.8))
    elements.append(Spacer(1, 0.35 * cm))

    global_df = pd.DataFrame({
        "N°": range(1, len(YEARS) + 1),
        "Année": YEARS,
        "CA cible": [float(ca_dict[y]) for y in YEARS],
        "CA généré": [float(all_results[y]["CA Total"].sum()) for y in YEARS],
        "Écart": [float(all_results[y]["CA Total"].sum()) - float(ca_dict[y]) for y in YEARS]
    })
    elements.append(Paragraph("2. Contrôle global", sub_style))
    elements.append(build_pdf_table(global_df, header_color="#17375F", body_color="#F8F5EF", font_size=8.8))
    elements.append(PageBreak())

    for i, y in enumerate(YEARS, start=1):
        detail_df = add_line_numbers(all_results[y]).copy()
        monthly_df = add_line_numbers(all_monthly[y]).copy()

        elements.append(Paragraph(f"{y} - Synthèse", sub_style))
        recap = (
            f"CA cible : <b>{money(ca_dict[y])}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"CA généré : <b>{money(detail_df['CA Total'].sum())}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Quantité totale : <b>{money(detail_df['Qté Totale'].sum())}</b>"
        )
        elements.append(Paragraph(recap, normal_style))
        elements.append(Spacer(1, 0.15 * cm))

        monthly_table = build_pdf_table(
            monthly_df,
            header_color="#123A64",
            body_color="#F7F4EF",
            font_size=8.4,
            col_widths=[1.0 * cm, 5.6 * cm, 4.0 * cm, 4.5 * cm]
        )
        elements.append(Paragraph("Synthèse mensuelle", small_style))
        elements.append(monthly_table)
        elements.append(Spacer(1, 0.25 * cm))

        detail_cols = ["N°", "Acte / Examen", "Prix Unitaire"] + [f"Qté {m}" for m in MONTHS] + ["Qté Totale", "CA Total"]
        detail_pdf_df = detail_df[detail_cols].copy()
        chunks = split_dataframe_for_pdf(detail_pdf_df, max_rows=12)

        for idx_chunk, chunk in enumerate(chunks, start=1):
            if idx_chunk > 1:
                elements.append(PageBreak())
                elements.append(Paragraph(f"{y} - Détail complet (suite {idx_chunk})", sub_style))
            else:
                elements.append(Paragraph("Détail complet des lignes", small_style))

            col_widths = [0.8 * cm, 4.6 * cm, 1.7 * cm] + [1.1 * cm] * 12 + [1.5 * cm, 2.0 * cm]
            detail_table = build_pdf_table(
                chunk,
                header_color="#C99A67",
                body_color="#FBF7F2",
                font_size=6.7,
                col_widths=col_widths
            )
            elements.append(detail_table)
            elements.append(Spacer(1, 0.2 * cm))

        if i < len(YEARS):
            elements.append(PageBreak())

    doc.build(elements)
    output.seek(0)
    return output

# =========================================================
# SESSION
# =========================================================
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    login_page()
    st.stop()

try:
    init_db()
except Exception as e:
    st.error(f"Erreur base de données : {e}")
    st.stop()

# =========================================================
# HEADER
# =========================================================
st.markdown("""
<div class="hero">
    <h1>🏥 Génération des Quantités</h1>
    <p>
        <span class="badge">Sauvegarde durable</span>
        <span class="badge">Design sombre</span>
        <span class="badge">Excel premium</span>
        <span class="badge">PDF structuré</span>
    </p>
    <p>
       Chiffres d'affaires detaillés
    </p>
</div>
""", unsafe_allow_html=True)

m1, m2, m3, m4 = st.columns(4)
with m1:
    st.markdown('<div class="metric-box"><div class="metric-label">Rubriques</div><div class="metric-value">3</div></div>', unsafe_allow_html=True)
with m2:
    st.markdown('<div class="metric-box"><div class="metric-label">Années gérées</div><div class="metric-value">5</div></div>', unsafe_allow_html=True)
with m3:
    st.markdown('<div class="metric-box"><div class="metric-label">Persistance</div><div class="metric-value">Neon</div></div>', unsafe_allow_html=True)
with m4:
    st.markdown('<div class="metric-box"><div class="metric-label">Exports</div><div class="metric-value">Excel + PDF</div></div>', unsafe_allow_html=True)

st.write("")

# =========================================================
# SIDEBAR
# =========================================================
st.sidebar.title("Navigation")
selected_section = st.sidebar.radio(
    "Choisir une rubrique",
    ["Clinique", "Laboratoire", "Centre de radiologie"]
)

allow_decimals = st.sidebar.checkbox("Autoriser parfois des quantités décimales", value=False)
seed_value = st.sidebar.number_input("Seed aléatoire", min_value=0, value=42, step=1)

if st.sidebar.button("Se déconnecter", use_container_width=True):
    st.session_state["authenticated"] = False
    st.rerun()

section_items = load_section_items(selected_section)
section_ca = load_section_ca(selected_section)

# =========================================================
# TABS
# =========================================================
tab1, tab2, tab3, tab4 = st.tabs([
    "1. Paramétrage",
    "2. Génération",
    "3. Résultats",
    "4. Export"
])

with tab1:
    left, right = st.columns([0.95, 1.7], gap="large")

    with left:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">CA cible - {selected_section}</div>', unsafe_allow_html=True)
        st.markdown('<div class="small-text">Modifie les objectifs puis clique sur Enregistrer.</div>', unsafe_allow_html=True)

        edited_ca = {}
        for y in YEARS:
            edited_ca[y] = st.number_input(
                y,
                min_value=0.0,
                value=float(section_ca[y]),
                step=100000.0,
                key=f"{selected_section}_{y}"
            )
        st.markdown('</div>', unsafe_allow_html=True)

    with right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">Actes / Examens - {selected_section}</div>', unsafe_allow_html=True)
        st.markdown('<div class="small-text">Tu peux ajouter, modifier ou supprimer des lignes.</div>', unsafe_allow_html=True)

        edited_df = st.data_editor(
            add_line_numbers(section_items),
            num_rows="dynamic",
            use_container_width=True,
            disabled=["N°"],
            key=f"editor_{selected_section}"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    st.write("")
    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("💾 Enregistrer les modifications", use_container_width=True):
            try:
                to_save = edited_df.drop(columns=["N°"], errors="ignore")
                to_save = clean_items_df(to_save)

                if to_save.empty:
                    st.error("Ajoute au moins une ligne avant d’enregistrer.")
                elif (to_save["Prix Unitaire"] <= 0).any():
                    st.error("Tous les prix unitaires doivent être > 0.")
                else:
                    save_section_items(selected_section, to_save)
                    save_section_ca(selected_section, edited_ca)
                    st.success(f"Les paramètres de '{selected_section}' ont bien été enregistrés.")
                    st.rerun()
            except Exception as e:
                st.error(f"Erreur pendant l’enregistrement : {e}")

    with c2:
        if st.button("🔄 Recharger les valeurs enregistrées", use_container_width=True):
            st.rerun()

    with c3:
        if st.button("♻️ Restaurer les valeurs par défaut", use_container_width=True):
            try:
                df_default = pd.DataFrame(DEFAULT_SECTIONS[selected_section], columns=["Acte / Examen", "Prix Unitaire"])
                save_section_items(selected_section, df_default)
                save_section_ca(selected_section, DEFAULT_CA.copy())
                st.success("Valeurs par défaut restaurées.")
                st.rerun()
            except Exception as e:
                st.error(f"Erreur pendant la restauration : {e}")

with tab2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(f'<div class="section-title">Génération - {selected_section}</div>', unsafe_allow_html=True)
    st.markdown('<div class="small-text">Répartition métier plus réaliste + ajustement exact pour respecter le CA annuel.</div>', unsafe_allow_html=True)

    if st.button("⚙️ Générer les quantités", type="primary", use_container_width=True):
        try:
            items_df = load_section_items(selected_section)
            ca_dict = load_section_ca(selected_section)

            if items_df.empty:
                st.error("Aucune ligne enregistrée.")
            elif (items_df["Prix Unitaire"] <= 0).any():
                st.error("Tous les prix unitaires doivent être supérieurs à 0.")
            else:
                all_results = {}
                all_monthly = {}
                controls = []

                for idx, y in enumerate(YEARS):
                    detail_df, total_generated = generate_realistic_year(
                        section_name=selected_section,
                        items_df=items_df,
                        ca_target=ca_dict[y],
                        seed=seed_value + idx,
                        allow_decimals=allow_decimals
                    )
                    monthly_df = build_monthly_summary(detail_df)

                    all_results[y] = detail_df
                    all_monthly[y] = monthly_df
                    controls.append({
                        "Année": y,
                        "CA cible": float(ca_dict[y]),
                        "CA généré": float(total_generated),
                        "Écart": float(total_generated - ca_dict[y]),
                    })

                st.session_state[f"results_{selected_section}"] = all_results
                st.session_state[f"monthly_{selected_section}"] = all_monthly
                st.session_state[f"controls_{selected_section}"] = pd.DataFrame(controls)
                st.success("Génération terminée avec succès.")
        except Exception as e:
            st.error(f"Erreur pendant la génération : {e}")

    st.markdown('</div>', unsafe_allow_html=True)

with tab3:
    results_key = f"results_{selected_section}"
    monthly_key = f"monthly_{selected_section}"
    controls_key = f"controls_{selected_section}"

    if results_key not in st.session_state:
        st.info("Génère d’abord les résultats.")
    else:
        controls_df = add_line_numbers(format_df_display(st.session_state[controls_key]))
        all_results = st.session_state[results_key]
        all_monthly = st.session_state[monthly_key]

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Contrôle global</div>', unsafe_allow_html=True)
        st.dataframe(controls_df, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.write("")
        selected_year = st.radio("Choisir une année", YEARS, horizontal=True, key=f"year_{selected_section}")

        detail_df = add_line_numbers(all_results[selected_year])
        monthly_df = add_line_numbers(all_monthly[selected_year])

        a, b, c = st.columns(3)
        with a:
            st.markdown(f'<div class="metric-box"><div class="metric-label">Rubrique</div><div class="metric-value">{selected_section}</div></div>', unsafe_allow_html=True)
        with b:
            st.markdown(f'<div class="metric-box"><div class="metric-label">Année</div><div class="metric-value">{selected_year}</div></div>', unsafe_allow_html=True)
        with c:
            st.markdown(f'<div class="metric-box"><div class="metric-label">CA recalculé</div><div class="metric-value">{money(monthly_df["CA Mensuel"].sum())}</div></div>', unsafe_allow_html=True)

        st.write("")

        c1, c2 = st.columns([1.15, 1], gap="large")
        with c1:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">CA mensuel</div>', unsafe_allow_html=True)
            fig1 = px.bar(monthly_df, x="Mois", y="CA Mensuel", text_auto=".2s")
            fig1.update_layout(
                height=410,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E7EEF8"),
                xaxis=dict(showgrid=False),
                yaxis=dict(gridcolor="rgba(255,255,255,0.10)")
            )
            st.plotly_chart(fig1, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        with c2:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">CA annuel par acte / examen</div>', unsafe_allow_html=True)
            chart_df = detail_df[["Acte / Examen", "CA Total"]].sort_values("CA Total", ascending=False)
            fig2 = px.bar(chart_df, x="Acte / Examen", y="CA Total")
            fig2.update_layout(
                height=410,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E7EEF8"),
                xaxis=dict(showgrid=False),
                yaxis=dict(gridcolor="rgba(255,255,255,0.10)")
            )
            st.plotly_chart(fig2, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        st.write("")
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Synthèse mensuelle</div>', unsafe_allow_html=True)
        st.dataframe(format_df_display(monthly_df), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.write("")
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Détail complet</div>', unsafe_allow_html=True)
        st.dataframe(format_df_display(detail_df), use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

with tab4:
    results_key = f"results_{selected_section}"
    monthly_key = f"monthly_{selected_section}"

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Export premium</div>', unsafe_allow_html=True)
    st.markdown('<div class="small-text">Télécharge un Excel structuré et un PDF détaillé.</div>', unsafe_allow_html=True)

    if results_key not in st.session_state:
        st.info("Génère d’abord les résultats.")
    else:
        try:
            ca_dict = load_section_ca(selected_section)
            all_results = st.session_state[results_key]
            all_monthly = st.session_state[monthly_key]

            excel_file = export_excel_styled(selected_section, ca_dict, all_results, all_monthly)
            pdf_file = export_pdf(selected_section, ca_dict, all_results, all_monthly)

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "📥 Télécharger Excel premium",
                    data=excel_file,
                    file_name=f"{slugify_name(selected_section)}_quantites_premium.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with c2:
                st.download_button(
                    "📄 Télécharger PDF structuré",
                    data=pdf_file,
                    file_name=f"{slugify_name(selected_section)}_rapport_structure.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

            st.markdown(
                f"""
                <div class="kpi-note">
                    Export prêt pour <b>{selected_section}</b> : synthèse globale, synthèse mensuelle et détail complet.
                </div>
                """,
                unsafe_allow_html=True
            )
        except Exception as e:
            st.error(f"Erreur pendant l’export : {e}")

    st.markdown('</div>', unsafe_allow_html=True)